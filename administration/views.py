from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import TemplateView, ListView, CreateView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Q, Count, Avg, Sum
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.contrib.auth.models import User
from datetime import date, timedelta
from decimal import Decimal
import json

# Importar modelos
from authentication.models import UserProfile, UserGroup, GroupMembership
from academics.models import Grade, Subject, Course, AcademicYear, Student
from academics_extended.models import AcademicYear as ExtendedAcademicYear, Grade as ExtendedGrade, Subject as ExtendedSubject, Course as ExtendedCourse, SubjectAssignment, TimeSlot
# Schedule, GradeRecord, AttendanceRecord - TEMPORALMENTE DESHABILITADO


# Vista temporal de test para debuggear botones
class TestUsersView(LoginRequiredMixin, TemplateView):
    """Vista de test para debuggear botones de usuario"""
    template_name = 'administration/test_users.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['all_users'] = User.objects.select_related('profile').order_by('date_joined')
        return context


class AdminDashboardView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Dashboard principal para administradores"""
    template_name = 'administration/admin_dashboard.html'
    
    def test_func(self):
        return hasattr(self.request.user, 'profile') and self.request.user.profile.role == 'admin'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Estadísticas del sistema
        context.update({
            'total_students': Student.objects.count(),
            'total_teachers': User.objects.filter(profile__role='teacher').count(),
            'total_courses': ExtendedCourse.objects.count(),
            'total_subjects': ExtendedSubject.objects.count(),
            'total_grades': ExtendedGrade.objects.count(),
            'active_academic_year': ExtendedAcademicYear.objects.filter(is_current=True).first(),
        })
        
        return context


class AdminUserManagementView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista para gestión de usuarios del sistema (solo administradores)"""
    template_name = 'administration/admin_users.html'
    
    def test_func(self):
        return hasattr(self.request.user, 'profile') and self.request.user.profile.role == 'admin'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        from authentication.models import StudentEnrollment
        
        # Obtener todos los usuarios con sus perfiles
        users = User.objects.select_related('profile').all()
        
        # Obtener información de matrículas de estudiantes
        student_enrollments = StudentEnrollment.objects.select_related(
            'student', 'student__profile', 'homeroom_teacher'
        ).filter(academic_year='2025').order_by('grade', 'section', 'student__first_name')
        
        # Estadísticas de matrículas por grado
        enrollment_stats = {}
        for grade_choice in StudentEnrollment.GRADE_CHOICES:
            grade_code = grade_choice[0]
            grade_name = grade_choice[1]
            count = student_enrollments.filter(grade=grade_code).count()
            if count > 0:
                enrollment_stats[grade_name] = count
        
        # Estadísticas generales
        total_users = users.count()
        students = users.filter(profile__role='student')
        total_students = students.count()
        teachers = users.filter(profile__role='teacher')
        total_teachers = teachers.count()
        active_users = users.filter(is_active=True).count()
        
        # Datos para formularios
        grade_choices = StudentEnrollment.GRADE_CHOICES
        section_choices = StudentEnrollment.SECTION_CHOICES
        active_teachers = teachers.filter(is_active=True)
        
        # Estadísticas adicionales para el template original
        users_by_role = UserProfile.objects.values('role').annotate(count=Count('id'))
        recent_users = User.objects.filter(
            date_joined__gte=timezone.now() - timedelta(days=30)
        ).count()
        
        context.update({
            'all_users': users,  # Para mantener compatibilidad con el template original
            'users': users,
            'users_by_role': users_by_role,
            'recent_users': recent_users,
            'student_enrollments': student_enrollments,
            'enrollment_stats': enrollment_stats,
            'total_users': total_users,
            'total_students': total_students,
            'total_teachers': total_teachers,
            'active_users': active_users,
            'grade_choices': grade_choices,
            'section_choices': section_choices,
            'teachers': active_teachers,
            'students': students,
        })
        
        return context


class SystemConfigView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista para configuración del sistema (solo administradores)"""
    template_name = 'administration/system_config.html'
    
    def test_func(self):
        return hasattr(self.request.user, 'profile') and self.request.user.profile.role == 'admin'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Configuración académica
        context['academic_years'] = ExtendedAcademicYear.objects.all().order_by('-start_date')
        context['current_academic_year'] = ExtendedAcademicYear.objects.filter(is_current=True).first()
        
        # Estadísticas del sistema
        context['total_models'] = {
            'grades': ExtendedGrade.objects.count(),
            'subjects': ExtendedSubject.objects.count(),
            'courses': ExtendedCourse.objects.count(),
            'students': Student.objects.count(),
            'time_slots': TimeSlot.objects.count(),
        }
        
        return context


class BackupManagementView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista para gestión de respaldos (solo administradores)"""
    template_name = 'administration/backup_management.html'
    
    def test_func(self):
        return hasattr(self.request.user, 'profile') and self.request.user.profile.role == 'admin'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Información de la base de datos
        import os
        from django.conf import settings
        from datetime import datetime
        
        db_path = settings.DATABASES['default']['NAME']
        
        if os.path.exists(db_path):
            # Información del archivo de BD
            stat = os.stat(db_path)
            context['db_info'] = {
                'file_name': os.path.basename(db_path),
                'full_path': db_path,
                'size_mb': round(stat.st_size / 1024 / 1024, 2),
                'last_modified': datetime.fromtimestamp(stat.st_mtime),
                'table_count': self._get_table_count(),
            }
        else:
            context['db_info'] = {
                'file_name': 'No encontrado',
                'full_path': db_path,
                'size_mb': 0,
                'last_modified': None,
                'table_count': 0,
            }
        
        # Fecha actual para templates
        context['today'] = datetime.now()
        
        return context
    
    def _get_table_count(self):
        """Contar número de tablas en la base de datos"""
        try:
            from django.db import connection
            with connection.cursor() as cursor:
                cursor.execute("SELECT count(*) FROM sqlite_master WHERE type='table'")
                return cursor.fetchone()[0]
        except:
            return 0


class ProfileManagementView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista principal para gestión de perfiles de usuarios"""
    template_name = 'administration/profile_management.html'
    
    def test_func(self):
        return hasattr(self.request.user, 'profile') and self.request.user.profile.role == 'admin'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Todos los usuarios con sus perfiles
        all_users = User.objects.select_related('profile').order_by('date_joined')
        
        # Estadísticas de completitud de perfiles
        profile_stats = {
            'total_users': all_users.count(),
            'profiles_complete': all_users.filter(profile__is_profile_complete=True).count(),
            'profiles_incomplete': all_users.filter(profile__is_profile_complete=False).count(),
            'without_avatar': all_users.filter(profile__avatar='').count(),
        }
        
        # Usuarios por rol con estadísticas de completitud
        users_by_role = {}
        for role_code, role_name in UserProfile.ROLE_CHOICES:
            role_users = all_users.filter(profile__role=role_code)
            if role_users.exists():
                users_by_role[role_code] = {
                    'name': role_name,
                    'count': role_users.count(),
                    'complete': role_users.filter(profile__is_profile_complete=True).count(),
                    'users': role_users[:5]  # Primeros 5 para preview
                }
        
        # Perfiles con menor completitud (necesitan atención)
        incomplete_profiles = []
        for user in all_users.filter(profile__is_profile_complete=False)[:10]:
            if user.profile:
                incomplete_profiles.append({
                    'user': user,
                    'completion': user.profile.profile_completion_percentage,
                    'role': user.profile.get_role_display()
                })
        
        # Ordenar por menor completitud
        incomplete_profiles.sort(key=lambda x: x['completion'])
        
        context.update({
            'all_users': all_users,
            'profile_stats': profile_stats,
            'users_by_role': users_by_role,
            'incomplete_profiles': incomplete_profiles,
        })
        
        return context


class ProfileDetailView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista para ver/editar un perfil específico"""
    template_name = 'administration/profile_detail.html'
    
    def test_func(self):
        return hasattr(self.request.user, 'profile') and self.request.user.profile.role == 'admin'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        user_id = kwargs.get('user_id')
        try:
            user = User.objects.select_related('profile').get(id=user_id)
            profile = user.profile
            
            # Información del usuario y perfil
            context['target_user'] = user
            context['target_profile'] = profile
            
            # Formulario apropiado según el rol
            from authentication.forms import get_profile_form
            context['profile_form'] = get_profile_form(user, instance=profile)
            
            # Estadísticas del perfil
            context['profile_completion'] = profile.profile_completion_percentage
            context['is_complete'] = profile.is_profile_complete
            
            # Información adicional según el rol
            if profile.is_teacher:
                context['assigned_students'] = profile.get_assigned_students()
            elif profile.is_student:
                context['assigned_teacher'] = profile.get_assigned_teacher()
            
        except User.DoesNotExist:
            context['user_not_found'] = True
        
        return context


# === API VIEWS PARA GESTIÓN DE USUARIOS ===

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def toggle_user_status_api(request):
    """API para activar/desactivar usuarios"""
    if not (hasattr(request.user, 'profile') and request.user.profile.role == 'admin'):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        
        if not user_id:
            return JsonResponse({'error': 'ID de usuario requerido'}, status=400)
        
        user = User.objects.get(id=user_id)
        
        # No permitir desactivar al propio usuario
        if user == request.user:
            return JsonResponse({'error': 'No puedes desactivar tu propia cuenta'}, status=400)
        
        # Cambiar estado
        user.is_active = not user.is_active
        user.save()
        
        return JsonResponse({
            'success': True,
            'user_id': user.id,
            'is_active': user.is_active,
            'message': f'✅ Usuario "{user.username}" {"🟢 activado" if user.is_active else "🔴 desactivado"} exitosamente'
        })
        
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error: {str(e)}'}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def delete_user_api(request):
    """API para eliminar usuarios"""
    if not (hasattr(request.user, 'profile') and request.user.profile.role == 'admin'):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        
        if not user_id:
            return JsonResponse({'error': 'ID de usuario requerido'}, status=400)
        
        user = User.objects.get(id=user_id)
        
        # No permitir eliminar al propio usuario
        if user == request.user:
            return JsonResponse({'error': 'No puedes eliminar tu propia cuenta'}, status=400)
        
        # No permitir eliminar otros administradores
        if hasattr(user, 'profile') and user.profile.role == 'admin':
            return JsonResponse({'error': 'No se pueden eliminar otros administradores'}, status=400)
        
        username = user.username
        user.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'🗑️ Usuario "{username}" eliminado exitosamente del sistema'
        })
        
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error: {str(e)}'}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def create_user_api(request):
    """API para crear nuevos usuarios"""
    if not (hasattr(request.user, 'profile') and request.user.profile.role == 'admin'):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        # Verificar Content-Type
        if request.content_type != 'application/json':
            return JsonResponse({'error': 'Content-Type debe ser application/json'}, status=400)
        
        data = json.loads(request.body)
        
        # Validar datos requeridos
        required_fields = ['username', 'email', 'first_name', 'last_name', 'role', 'password']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({'error': f'Campo {field} es requerido'}, status=400)
        
        # Validar formato de username
        username = data['username'].strip()
        if not username.replace('_', '').replace('-', '').isalnum():
            return JsonResponse({'error': 'El username solo puede contener letras, números, guiones y guiones bajos'}, status=400)
        
        # Validar longitud de contraseña
        if len(data['password']) < 6:
            return JsonResponse({'error': 'La contraseña debe tener al menos 6 caracteres'}, status=400)
        
        # Verificar que el username no existe
        if User.objects.filter(username=username).exists():
            return JsonResponse({'error': 'El nombre de usuario ya existe'}, status=400)
        
        # Verificar que el email no existe
        email = data['email'].strip()
        if User.objects.filter(email=email).exists():
            return JsonResponse({'error': 'El email ya está en uso'}, status=400)
        
        # Validar rol
        valid_roles = ['admin', 'secretary', 'teacher']
        if data['role'] not in valid_roles:
            return JsonResponse({'error': f'Rol inválido. Use: {", ".join(valid_roles)}'}, status=400)
        
        # Crear usuario
        user = User.objects.create_user(
            username=username,
            email=email,
            password=data['password'],
            first_name=data['first_name'].strip(),
            last_name=data['last_name'].strip()
        )
        
        # Crear o actualizar perfil (la señal ya pudo haber creado uno con rol 'student')
        profile, created = UserProfile.objects.get_or_create(
            user=user,
            defaults={'role': data['role']}
        )
        
        # Si el perfil ya existía, actualizar el rol
        if not created:
            profile.role = data['role']
            profile.save()
        
        return JsonResponse({
            'success': True,
            'user_id': user.id,
            'message': f'🎉 ¡Usuario "{user.username}" creado exitosamente!\n👤 {user.first_name} {user.last_name}\n🎭 Rol: {profile.get_role_display()}\n📧 Email: {user.email}'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)
    except Exception as e:
        import traceback
        print(f"Error en create_user_api: {e}")
        print(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({'error': f'Error interno del servidor: {str(e)}'}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def reset_password_api(request):
    """API para resetear contraseña de usuarios"""
    if not (hasattr(request.user, 'profile') and request.user.profile.role == 'admin'):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        new_password = data.get('new_password', '123456')  # Contraseña temporal por defecto
        
        if not user_id:
            return JsonResponse({'error': 'ID de usuario requerido'}, status=400)
        
        user = User.objects.get(id=user_id)
        user.set_password(new_password)
        user.save()
        
        return JsonResponse({
            'success': True,
            'message': f'🔑 Contraseña de "{user.username}" reseteada exitosamente\n🎯 Nueva contraseña temporal generada',
            'temp_password': new_password
        })
        
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error: {str(e)}'}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def update_user_api(request):
    """API para actualizar información de usuarios"""
    if not (hasattr(request.user, 'profile') and request.user.profile.role == 'admin'):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        
        if not user_id:
            return JsonResponse({'error': 'ID de usuario requerido'}, status=400)
        
        user = User.objects.get(id=user_id)
        
        # Actualizar datos básicos del usuario
        if 'username' in data and data['username'] != user.username:
            # Verificar que el nuevo username no existe
            if User.objects.filter(username=data['username']).exclude(id=user_id).exists():
                return JsonResponse({'error': 'El nombre de usuario ya existe'}, status=400)
            user.username = data['username']
        
        if 'email' in data and data['email'] != user.email:
            # Verificar que el nuevo email no existe
            if User.objects.filter(email=data['email']).exclude(id=user_id).exists():
                return JsonResponse({'error': 'El email ya está en uso'}, status=400)
            user.email = data['email']
        
        if 'first_name' in data:
            user.first_name = data['first_name']
        
        if 'last_name' in data:
            user.last_name = data['last_name']
        
        user.save()
        
        # Actualizar rol si es necesario
        if 'role' in data and hasattr(user, 'profile'):
            user.profile.role = data['role']
            user.profile.save()
        
        return JsonResponse({
            'success': True,
            'message': f'📝 Usuario "{user.username}" actualizado exitosamente\n✨ Los cambios se han guardado correctamente'
        })
        
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error: {str(e)}'}, status=500)


# === API VIEWS PARA GESTIÓN DE PERFILES ===

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def update_profile_api(request):
    """API para actualizar perfil completo de usuario"""
    if not (hasattr(request.user, 'profile') and request.user.profile.role == 'admin'):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        
        if not user_id:
            return JsonResponse({'error': 'ID de usuario requerido'}, status=400)
        
        user = User.objects.get(id=user_id)
        profile = user.profile
        
        # Obtener el formulario apropiado para el rol
        from authentication.forms import get_profile_form
        form = get_profile_form(user, role=profile.role, data=data, instance=profile)
        
        if form.is_valid():
            # Guardar cambios en el perfil
            profile = form.save(commit=False)
            
            # Actualizar también los campos del User si están en el formulario
            if 'first_name' in form.cleaned_data:
                user.first_name = form.cleaned_data['first_name']
            if 'last_name' in form.cleaned_data:
                user.last_name = form.cleaned_data['last_name']
            if 'email' in form.cleaned_data:
                user.email = form.cleaned_data['email']
            
            user.save()
            profile.save()
            
            # Recalcular completitud del perfil
            profile.mark_profile_complete()
            
            return JsonResponse({
                'success': True,
                'message': f'✅ Perfil de "{user.username}" actualizado exitosamente\n📊 Completitud: {profile.profile_completion_percentage}%',
                'completion_percentage': profile.profile_completion_percentage,
                'is_complete': profile.is_profile_complete
            })
        else:
            # Devolver errores del formulario
            errors = []
            for field, error_list in form.errors.items():
                for error in error_list:
                    field_obj = form.fields.get(field)
                    field_label = getattr(field_obj, 'label', field) if field_obj else field
                    errors.append(f"{field_label}: {error}")
            
            return JsonResponse({
                'error': 'Errores de validación',
                'details': errors
            }, status=400)
        
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
    except Exception as e:
        import traceback
        print(f"Error en update_profile_api: {e}")
        print(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({'error': f'Error: {str(e)}'}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def upload_avatar_api(request):
    """API para subir avatar de usuario"""
    if not (hasattr(request.user, 'profile') and request.user.profile.role == 'admin'):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        user_id = request.POST.get('user_id')
        
        if not user_id:
            return JsonResponse({'error': 'ID de usuario requerido'}, status=400)
        
        user = User.objects.get(id=user_id)
        profile = user.profile
        
        # Verificar que se subió un archivo
        if 'avatar' not in request.FILES:
            return JsonResponse({'error': 'No se seleccionó ningún archivo'}, status=400)
        
        avatar_file = request.FILES['avatar']
        
        # Validar el archivo usando el formulario
        from authentication.forms import ProfileImageForm
        form = ProfileImageForm(data={}, files={'avatar': avatar_file}, instance=profile)
        
        if form.is_valid():
            profile = form.save()
            
            return JsonResponse({
                'success': True,
                'message': f'📸 Avatar de "{user.username}" actualizado exitosamente',
                'avatar_url': profile.avatar.url if profile.avatar else None
            })
        else:
            errors = []
            for field, error_list in form.errors.items():
                for error in error_list:
                    errors.append(error)
            
            return JsonResponse({
                'error': 'Error en el archivo',
                'details': errors
            }, status=400)
        
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error: {str(e)}'}, status=500)


@login_required
@require_http_methods(["GET"])
def profile_stats_api(request):
    """API para obtener estadísticas de perfiles"""
    if not (hasattr(request.user, 'profile') and request.user.profile.role == 'admin'):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        # Estadísticas generales
        total_users = User.objects.count()
        complete_profiles = UserProfile.objects.filter(is_profile_complete=True).count()
        with_avatars = UserProfile.objects.exclude(avatar='').count()
        
        # Estadísticas por rol
        role_stats = {}
        for role_code, role_name in UserProfile.ROLE_CHOICES:
            role_profiles = UserProfile.objects.filter(role=role_code)
            if role_profiles.exists():
                role_stats[role_code] = {
                    'name': role_name,
                    'total': role_profiles.count(),
                    'complete': role_profiles.filter(is_profile_complete=True).count(),
                    'with_avatar': role_profiles.exclude(avatar='').count(),
                }
        
        # Completitud promedio (calculada manualmente porque es una propiedad)
        all_profiles = UserProfile.objects.all()
        if all_profiles.exists():
            total_completion = sum(profile.profile_completion_percentage for profile in all_profiles)
            avg_completion = total_completion / all_profiles.count()
        else:
            avg_completion = 0
        
        return JsonResponse({
            'success': True,
            'stats': {
                'total_users': total_users,
                'complete_profiles': complete_profiles,
                'completion_rate': round((complete_profiles / total_users * 100), 1) if total_users > 0 else 0,
                'with_avatars': with_avatars,
                'avatar_rate': round((with_avatars / total_users * 100), 1) if total_users > 0 else 0,
                'avg_completion': round(avg_completion, 1),
                'role_stats': role_stats
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error: {str(e)}'}, status=500)


@login_required
@require_http_methods(["GET"])
def export_profiles_api(request):
    """API para exportar perfiles de usuario a Excel"""
    if not (hasattr(request.user, 'profile') and request.user.profile.role == 'admin'):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        from datetime import datetime
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Perfiles de Usuario"
        
        # Encabezados
        headers = [
            'ID', 'Nombre de Usuario', 'Nombre Completo', 'Email', 'Rol',
            'Teléfono', 'Celular', 'Ciudad', 'Dirección', 'Documento ID',
            'Fecha Nacimiento', 'Género', 'Título Profesional', 'Institución',
            'Año Graduación', 'Especialización', 'Departamento', 'Cargo',
            'Años Experiencia', 'Fecha Contratación', 'Estado', 'Avatar',
            'Completitud %', 'Perfil Completo', 'Contacto Emergencia',
            'Teléfono Emergencia', 'Fecha Registro', 'Última Actividad'
        ]
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Obtener datos de usuarios
        users = User.objects.select_related('profile').order_by('username')
        
        # Escribir datos
        for row, user in enumerate(users, 2):
            profile = user.profile
            
            data = [
                user.id,
                user.username,
                f"{user.first_name} {user.last_name}".strip(),
                user.email,
                profile.get_role_display(),
                profile.phone or '',
                profile.mobile_phone or '',
                profile.city or '',
                profile.address or '',
                profile.identification_number or '',
                profile.date_of_birth.strftime('%Y-%m-%d') if profile.date_of_birth else '',
                profile.get_gender_display() if profile.gender else '',
                profile.professional_title or '',
                profile.institution or '',
                profile.graduation_year or '',
                profile.specialization or '',
                profile.department or '',
                profile.position or '',
                profile.years_of_experience or '',
                profile.hire_date.strftime('%Y-%m-%d') if profile.hire_date else '',
                'Activo' if user.is_active else 'Inactivo',
                'Sí' if profile.avatar else 'No',
                f"{profile.profile_completion_percentage}%",
                'Sí' if profile.is_profile_complete else 'No',
                profile.emergency_contact_name or '',
                profile.emergency_contact_phone or '',
                user.date_joined.strftime('%Y-%m-%d %H:%M'),
                user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Nunca'
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'perfiles_usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar workbook en respuesta
        wb.save(response)
        
        return response
        
    except ImportError:
        return JsonResponse({
            'error': 'La librería openpyxl no está instalada. Instale con: pip install openpyxl'
        }, status=500)
    except Exception as e:
        return JsonResponse({'error': f'Error al exportar: {str(e)}'}, status=500)


# =============================================================================
# GESTIÓN DE GRUPOS DE USUARIOS
# =============================================================================

class GroupManagementView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista principal para gestión de grupos/cursos y asignación de directores de grupo"""
    template_name = 'administration/group_management.html'
    
    def test_func(self):
        return (hasattr(self.request.user, 'profile') and 
                self.request.user.profile.role == 'admin')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Importar modelos académicos
        from academics_extended.models import AcademicYear, Grade, Course
        
        # Año académico actual
        try:
            current_year = AcademicYear.objects.filter(is_current=True).first()
            if not current_year:
                # Si no hay año actual, crear uno por defecto
                from datetime import date
                current_year = AcademicYear.objects.create(
                    name="2025",
                    start_date=date(2025, 1, 1),
                    end_date=date(2025, 12, 31),
                    is_current=True
                )
        except:
            current_year = None
        
        # Obtener todos los cursos del año académico actual
        if current_year:
            courses = Course.objects.filter(
                academic_year=current_year
            ).select_related(
                'grade', 'homeroom_teacher', 'homeroom_teacher__profile'
            ).order_by('grade__order', 'section')
        else:
            courses = Course.objects.none()
        
        # Profesores disponibles para asignar como directores de grupo
        available_teachers = User.objects.filter(
            is_active=True, 
            profile__role='teacher'
        ).select_related('profile').order_by('first_name', 'last_name')
        
        # Grados disponibles
        grades = Grade.objects.all().order_by('order')
        
        # Estadísticas
        total_courses = courses.count()
        courses_with_teacher = courses.filter(homeroom_teacher__isnull=False).count()
        courses_without_teacher = total_courses - courses_with_teacher
        total_teachers = available_teachers.count()
        
        # Cursos agrupados por grado para mejor visualización
        courses_by_grade = {}
        for course in courses:
            grade_name = course.grade.name
            if grade_name not in courses_by_grade:
                courses_by_grade[grade_name] = {
                    'grade': course.grade,
                    'courses': []
                }
            courses_by_grade[grade_name]['courses'].append(course)
        
        # Profesores que ya son directores de grupo
        assigned_teachers = []
        unassigned_teachers = []
        
        for teacher in available_teachers:
            teacher_courses = courses.filter(homeroom_teacher=teacher)
            if teacher_courses.exists():
                assigned_teachers.append({
                    'teacher': teacher,
                    'courses': list(teacher_courses),
                    'course_count': teacher_courses.count()
                })
            else:
                unassigned_teachers.append(teacher)
        
        context.update({
            # Datos principales
            'current_year': current_year,
            'courses': courses,
            'grades': grades,
            'available_teachers': available_teachers,
            'assigned_teachers': assigned_teachers,
            'unassigned_teachers': unassigned_teachers,
            
            # Estadísticas
            'total_courses': total_courses,
            'courses_with_teacher': courses_with_teacher,
            'courses_without_teacher': courses_without_teacher,
            'total_teachers': total_teachers,
            'assigned_teachers_count': len(assigned_teachers),
            'unassigned_teachers_count': len(unassigned_teachers),
            
            # Vista organizada
            'courses_by_grade': courses_by_grade,
        })
        
        return context


# === API ENDPOINTS PARA GESTIÓN DE DIRECTORES DE GRUPO ===

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def assign_teacher_to_course_api(request):
    """API para asignar un profesor como director de grupo"""
    if not (hasattr(request.user, 'profile') and request.user.profile.role == 'admin'):
        return JsonResponse({'error': 'Sin permisos de administrador'}, status=403)
    
    try:
        data = json.loads(request.body)
        course_id = data.get('course_id')
        teacher_id = data.get('teacher_id')
        
        if not course_id or not teacher_id:
            return JsonResponse({'error': 'ID de curso y profesor requeridos'}, status=400)
        
        # Importar modelos académicos
        from academics_extended.models import Course
        
        # Obtener el curso
        course = Course.objects.get(id=course_id)
        
        # Obtener el profesor
        teacher = User.objects.get(id=teacher_id, profile__role='teacher', is_active=True)
        
        # Verificar si el profesor ya es director de otro curso
        existing_assignment = Course.objects.filter(homeroom_teacher=teacher).first()
        if existing_assignment and existing_assignment.id != course.id:
            return JsonResponse({
                'error': f'El profesor ya es director del curso {existing_assignment.grade.name} {existing_assignment.section}'
            }, status=400)
        
        # Guardar asignación anterior para el mensaje
        previous_teacher = course.homeroom_teacher
        
        # Asignar el profesor al curso
        course.homeroom_teacher = teacher
        course.save()
        
        # Preparar mensaje de respuesta
        if previous_teacher:
            message = f'Director cambiado: {teacher.get_full_name()} ahora es el director de {course.grade.name} {course.section} (reemplazando a {previous_teacher.get_full_name()})'
        else:
            message = f'Director asignado: {teacher.get_full_name()} es ahora el director de {course.grade.name} {course.section}'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'course_id': course.id,
            'teacher_id': teacher.id,
            'teacher_name': teacher.get_full_name(),
            'course_name': f'{course.grade.name} {course.section}'
        })
        
    except Course.DoesNotExist:
        return JsonResponse({'error': 'Curso no encontrado'}, status=404)
    except User.DoesNotExist:
        return JsonResponse({'error': 'Profesor no encontrado o inactivo'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)


@login_required
@csrf_exempt  
@require_http_methods(["POST"])
def remove_teacher_from_course_api(request):
    """API para quitar un profesor como director de grupo"""
    if not (hasattr(request.user, 'profile') and request.user.profile.role == 'admin'):
        return JsonResponse({'error': 'Sin permisos de administrador'}, status=403)
    
    try:
        data = json.loads(request.body)
        course_id = data.get('course_id')
        
        if not course_id:
            return JsonResponse({'error': 'ID de curso requerido'}, status=400)
        
        # Importar modelos académicos
        from academics_extended.models import Course
        
        # Obtener el curso
        course = Course.objects.get(id=course_id)
        
        if not course.homeroom_teacher:
            return JsonResponse({'error': 'Este curso no tiene director asignado'}, status=400)
        
        # Guardar información del profesor para el mensaje
        removed_teacher = course.homeroom_teacher
        
        # Quitar el profesor del curso
        course.homeroom_teacher = None
        course.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Director removido: {removed_teacher.get_full_name()} ya no es director de {course.grade.name} {course.section}',
            'course_id': course.id,
            'course_name': f'{course.grade.name} {course.section}',
            'removed_teacher': removed_teacher.get_full_name()
        })
        
    except Course.DoesNotExist:
        return JsonResponse({'error': 'Curso no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)


class GroupDetailView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista detallada de un grupo específico"""
    template_name = 'administration/group_detail.html'
    
    def test_func(self):
        return (hasattr(self.request.user, 'profile') and 
                self.request.user.profile.role == 'admin')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        group_id = kwargs.get('group_id')
        
        try:
            group = UserGroup.objects.select_related('created_by').prefetch_related(
                'memberships__user__profile'
            ).get(id=group_id)
            
            context.update({
                'group': group,
                'members': group.memberships.filter(is_active=True).select_related('user__profile'),
                'leaders': group.memberships.filter(is_active=True, role='leader').select_related('user__profile'),
                'available_users': User.objects.filter(is_active=True).exclude(
                    id__in=group.memberships.filter(is_active=True).values_list('user_id', flat=True)
                ).order_by('first_name', 'last_name')
            })
            
        except UserGroup.DoesNotExist:
            context['group_not_found'] = True
            
        return context


# =============================================================================
# APIs PARA GESTIÓN DE GRUPOS
# =============================================================================

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def create_group_api(request):
    """API para crear un nuevo grupo"""
    if not (hasattr(request.user, 'profile') and request.user.profile.role == 'admin'):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        # Validar datos requeridos
        name = data.get('name', '').strip()
        if not name:
            return JsonResponse({'error': 'El nombre del grupo es requerido'}, status=400)
        
        # Verificar que no exista un grupo con el mismo nombre
        if UserGroup.objects.filter(name=name).exists():
            return JsonResponse({'error': 'Ya existe un grupo con ese nombre'}, status=400)
        
        # Crear el grupo
        group = UserGroup.objects.create(
            name=name,
            description=data.get('description', ''),
            group_type=data.get('group_type', 'functional'),
            created_by=request.user,
            allow_self_join=data.get('allow_self_join', False),
            is_public=data.get('is_public', True)
        )
        
        # Si se especificó un profesor monitor, agregarlo como líder
        monitor_id = data.get('monitor_id')
        if monitor_id:
            try:
                monitor = User.objects.get(id=monitor_id, profile__role='teacher')
                GroupMembership.objects.create(
                    user=monitor,
                    group=group,
                    role='leader',
                    added_by=request.user
                )
            except User.DoesNotExist:
                pass  # Continuar aunque no se encuentre el monitor
        
        # Agregar el creador como líder si no hay monitor o si el admin también debe estar
        if not monitor_id or request.user.profile.role == 'admin':
            GroupMembership.objects.get_or_create(
                user=request.user,
                group=group,
                defaults={
                    'role': 'leader' if not monitor_id else 'moderator',
                    'added_by': request.user
                }
            )
        
        # Agregar miembros iniciales si se especificaron
        initial_members = data.get('initial_members', [])
        if initial_members:
            for member_id in initial_members:
                try:
                    member = User.objects.get(id=member_id, is_active=True)
                    # Evitar duplicados
                    if not GroupMembership.objects.filter(user=member, group=group).exists():
                        role = 'moderator' if member.profile.role == 'teacher' else 'member'
                        GroupMembership.objects.create(
                            user=member,
                            group=group,
                            role=role,
                            added_by=request.user
                        )
                except User.DoesNotExist:
                    continue  # Continuar con el siguiente miembro
        
        return JsonResponse({
            'success': True,
            'message': f'✅ Grupo "{name}" creado exitosamente',
            'group_id': group.id
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error al crear grupo: {str(e)}'}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def update_group_api(request):
    """API para actualizar un grupo"""
    if not (hasattr(request.user, 'profile') and request.user.profile.role == 'admin'):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        data = json.loads(request.body)
        group_id = data.get('group_id')
        
        if not group_id:
            return JsonResponse({'error': 'ID de grupo requerido'}, status=400)
        
        group = UserGroup.objects.get(id=group_id)
        
        # Actualizar campos
        if 'name' in data:
            name = data['name'].strip()
            if not name:
                return JsonResponse({'error': 'El nombre no puede estar vacío'}, status=400)
            # Verificar nombre único (excluyendo el grupo actual)
            if UserGroup.objects.filter(name=name).exclude(id=group_id).exists():
                return JsonResponse({'error': 'Ya existe otro grupo con ese nombre'}, status=400)
            group.name = name
        
        if 'description' in data:
            group.description = data['description']
        if 'group_type' in data:
            group.group_type = data['group_type']
        if 'allow_self_join' in data:
            group.allow_self_join = data['allow_self_join']
        if 'is_public' in data:
            group.is_public = data['is_public']
        if 'is_active' in data:
            group.is_active = data['is_active']
        
        group.save()
        
        return JsonResponse({
            'success': True,
            'message': f'✅ Grupo "{group.name}" actualizado exitosamente'
        })
        
    except UserGroup.DoesNotExist:
        return JsonResponse({'error': 'Grupo no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error al actualizar grupo: {str(e)}'}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def delete_group_api(request):
    """API para eliminar un grupo"""
    if not (hasattr(request.user, 'profile') and request.user.profile.role == 'admin'):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        data = json.loads(request.body)
        group_id = data.get('group_id')
        
        if not group_id:
            return JsonResponse({'error': 'ID de grupo requerido'}, status=400)
        
        group = UserGroup.objects.get(id=group_id)
        group_name = group.name
        
        # Eliminar todas las membresías primero
        GroupMembership.objects.filter(group=group).delete()
        
        # Eliminar el grupo
        group.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'✅ Grupo "{group_name}" eliminado exitosamente'
        })
        
    except UserGroup.DoesNotExist:
        return JsonResponse({'error': 'Grupo no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error al eliminar grupo: {str(e)}'}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def add_group_member_api(request):
    """API para agregar un miembro a un grupo"""
    if not (hasattr(request.user, 'profile') and request.user.profile.role == 'admin'):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        data = json.loads(request.body)
        group_id = data.get('group_id')
        user_id = data.get('user_id')
        role = data.get('role', 'member')
        
        if not group_id or not user_id:
            return JsonResponse({'error': 'ID de grupo y usuario requeridos'}, status=400)
        
        group = UserGroup.objects.get(id=group_id)
        user = User.objects.get(id=user_id)
        
        # Verificar que el usuario no esté ya en el grupo
        if GroupMembership.objects.filter(group=group, user=user, is_active=True).exists():
            return JsonResponse({'error': f'{user.get_full_name()} ya es miembro del grupo'}, status=400)
        
        # Crear o reactivar la membresía
        membership, created = GroupMembership.objects.get_or_create(
            group=group,
            user=user,
            defaults={
                'role': role,
                'added_by': request.user,
                'is_active': True
            }
        )
        
        if not created:
            # Reactivar membresía existente
            membership.role = role
            membership.is_active = True
            membership.added_by = request.user
            membership.save()
        
        return JsonResponse({
            'success': True,
            'message': f'✅ {user.get_full_name()} agregado al grupo "{group.name}" como {membership.get_role_display()}'
        })
        
    except UserGroup.DoesNotExist:
        return JsonResponse({'error': 'Grupo no encontrado'}, status=404)
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error al agregar miembro: {str(e)}'}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def remove_group_member_api(request):
    """API para remover un miembro de un grupo"""
    if not (hasattr(request.user, 'profile') and request.user.profile.role == 'admin'):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        data = json.loads(request.body)
        group_id = data.get('group_id')
        user_id = data.get('user_id')
        
        if not group_id or not user_id:
            return JsonResponse({'error': 'ID de grupo y usuario requeridos'}, status=400)
        
        group = UserGroup.objects.get(id=group_id)
        user = User.objects.get(id=user_id)
        
        # Buscar la membresía activa
        membership = GroupMembership.objects.filter(
            group=group, 
            user=user, 
            is_active=True
        ).first()
        
        if not membership:
            return JsonResponse({'error': f'{user.get_full_name()} no es miembro activo del grupo'}, status=400)
        
        # Desactivar la membresía (no eliminar para mantener historial)
        membership.is_active = False
        membership.save()
        
        return JsonResponse({
            'success': True,
            'message': f'✅ {user.get_full_name()} removido del grupo "{group.name}"'
        })
        
    except UserGroup.DoesNotExist:
        return JsonResponse({'error': 'Grupo no encontrado'}, status=404)
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error al remover miembro: {str(e)}'}, status=500)


# =============================================================================
# GESTIÓN DE ESTUDIANTES Y MATRÍCULAS
# =============================================================================

class StudentManagementView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista principal para gestión de estudiantes y matrículas"""
    template_name = 'administration/student_management.html'
    
    def test_func(self):
        return (hasattr(self.request.user, 'profile') and 
                self.request.user.profile.role in ['admin', 'secretary'])
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        from authentication.models import StudentEnrollment
        
        # Obtener matrículas del año actual
        current_year = '2025'
        enrollments = StudentEnrollment.objects.select_related(
            'student', 'student__profile', 'homeroom_teacher', 'homeroom_teacher__profile'
        ).filter(academic_year=current_year).order_by('grade', 'section', 'student__last_name')
        
        # Estadísticas generales
        total_students = enrollments.count()
        active_students = enrollments.filter(status='active').count()
        enrolled_students = enrollments.filter(status='enrolled').count()
        
        # Estadísticas por grado
        primary_students = enrollments.filter(grade__in=['preescolar', 'primero', 'segundo', 'tercero', 'cuarto', 'quinto']).count()
        secondary_students = enrollments.filter(grade__in=['sexto', 'septimo', 'octavo', 'noveno', 'decimo', 'once']).count()
        
        # Profesores disponibles para director de grupo
        teachers = User.objects.filter(
            is_active=True, 
            profile__role='teacher'
        ).select_related('profile').order_by('first_name', 'last_name')
        
        context.update({
            'enrollments': enrollments,
            'total_students': total_students,
            'active_students': active_students,
            'enrolled_students': enrolled_students,
            'primary_students': primary_students,
            'secondary_students': secondary_students,
            'teachers': teachers,
            'current_year': current_year,
            'grade_choices': StudentEnrollment.GRADE_CHOICES,
            'section_choices': StudentEnrollment.SECTION_CHOICES,
            'status_choices': StudentEnrollment.STATUS_CHOICES,
        })
        
        return context


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def create_student_enrollment_api(request):
    """API para crear una nueva matrícula de estudiante"""
    if not (hasattr(request.user, 'profile') and 
            request.user.profile.role in ['admin', 'secretary']):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        from authentication.models import StudentEnrollment
        data = json.loads(request.body)
        
        # Datos del estudiante
        student_data = {
            'username': data.get('username', '').strip(),
            'email': data.get('email', '').strip(),
            'first_name': data.get('first_name', '').strip(),
            'last_name': data.get('last_name', '').strip(),
        }
        
        # Validar datos básicos
        if not all([student_data['username'], student_data['first_name'], student_data['last_name']]):
            return JsonResponse({'error': 'Username, nombre y apellido son requeridos'}, status=400)
        
        # Verificar que no exista el usuario
        if User.objects.filter(username=student_data['username']).exists():
            return JsonResponse({'error': 'Ya existe un usuario con ese username'}, status=400)
        
        # Datos de matrícula
        enrollment_data = {
            'academic_year': data.get('academic_year', '2025'),
            'grade': data.get('grade'),
            'section': data.get('section', 'A'),
            'homeroom_teacher_id': data.get('homeroom_teacher_id'),
            'parent_guardian_name': data.get('parent_guardian_name', '').strip(),
            'parent_guardian_phone': data.get('parent_guardian_phone', '').strip(),
            'parent_guardian_email': data.get('parent_guardian_email', '').strip(),
        }
        
        # Validar datos de matrícula
        if not enrollment_data['grade']:
            return JsonResponse({'error': 'El grado es requerido'}, status=400)
        
        if not enrollment_data['parent_guardian_name']:
            return JsonResponse({'error': 'El nombre del acudiente es requerido'}, status=400)
        
        # Verificar que no exista otra matrícula para el mismo año
        existing_enrollment = StudentEnrollment.objects.filter(
            student__username=student_data['username'],
            academic_year=enrollment_data['academic_year']
        ).first()
        
        if existing_enrollment:
            return JsonResponse({'error': 'Ya existe una matrícula para este estudiante en el año académico'}, status=400)
        
        with transaction.atomic():
            # Crear el usuario estudiante
            student = User.objects.create_user(
                username=student_data['username'],
                email=student_data['email'],
                first_name=student_data['first_name'],
                last_name=student_data['last_name'],
                password='estudiante123'  # Contraseña por defecto
            )
            
            # Configurar perfil
            student.profile.role = 'student'
            student.profile.phone = enrollment_data['parent_guardian_phone']
            student.profile.save()
            
            # Obtener profesor director si se especificó
            homeroom_teacher = None
            if enrollment_data['homeroom_teacher_id']:
                try:
                    homeroom_teacher = User.objects.get(
                        id=enrollment_data['homeroom_teacher_id'],
                        profile__role='teacher'
                    )
                except User.DoesNotExist:
                    pass
            
            # Crear la matrícula
            enrollment = StudentEnrollment.objects.create(
                student=student,
                academic_year=enrollment_data['academic_year'],
                grade=enrollment_data['grade'],
                section=enrollment_data['section'],
                homeroom_teacher=homeroom_teacher,
                parent_guardian_name=enrollment_data['parent_guardian_name'],
                parent_guardian_phone=enrollment_data['parent_guardian_phone'],
                parent_guardian_email=enrollment_data['parent_guardian_email'],
                status='enrolled',
                created_by=request.user
            )
            
            # Asignar automáticamente al grupo académico
            enrollment.assign_to_academic_group()
            
            return JsonResponse({
                'success': True,
                'message': f'✅ Estudiante {student.get_full_name()} matriculado exitosamente en {enrollment.full_grade}'
            })
        
    except Exception as e:
        return JsonResponse({'error': f'Error al crear matrícula: {str(e)}'}, status=500)


@login_required
@csrf_exempt  
@require_http_methods(["POST"])
def update_student_enrollment_api(request):
    """API para actualizar una matrícula de estudiante"""
    if not (hasattr(request.user, 'profile') and 
            request.user.profile.role in ['admin', 'secretary']):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        from authentication.models import StudentEnrollment
        data = json.loads(request.body)
        enrollment_id = data.get('enrollment_id')
        
        if not enrollment_id:
            return JsonResponse({'error': 'ID de matrícula requerido'}, status=400)
        
        enrollment = StudentEnrollment.objects.get(id=enrollment_id)
        
        # Actualizar datos del estudiante
        student = enrollment.student
        if 'first_name' in data and data['first_name'].strip():
            student.first_name = data['first_name'].strip()
        if 'last_name' in data and data['last_name'].strip():
            student.last_name = data['last_name'].strip()
        if 'email' in data:
            student.email = data['email'].strip()
        student.save()
        
        # Actualizar datos de matrícula
        if 'grade' in data:
            enrollment.grade = data['grade']
        if 'section' in data:
            enrollment.section = data['section']
        if 'status' in data:
            enrollment.status = data['status']
        if 'homeroom_teacher_id' in data:
            if data['homeroom_teacher_id']:
                try:
                    teacher = User.objects.get(id=data['homeroom_teacher_id'], profile__role='teacher')
                    enrollment.homeroom_teacher = teacher
                except User.DoesNotExist:
                    pass
            else:
                enrollment.homeroom_teacher = None
        
        # Actualizar datos del acudiente
        if 'parent_guardian_name' in data:
            enrollment.parent_guardian_name = data['parent_guardian_name']
        if 'parent_guardian_phone' in data:
            enrollment.parent_guardian_phone = data['parent_guardian_phone']
        if 'parent_guardian_email' in data:
            enrollment.parent_guardian_email = data['parent_guardian_email']
        
        enrollment.save()
        
        return JsonResponse({
            'success': True,
            'message': f'✅ Matrícula de {student.get_full_name()} actualizada exitosamente'
        })
        
    except StudentEnrollment.DoesNotExist:
        return JsonResponse({'error': 'Matrícula no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error al actualizar matrícula: {str(e)}'}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def change_student_status_api(request):
    """API para cambiar el estado de un estudiante"""
    if not (hasattr(request.user, 'profile') and 
            request.user.profile.role in ['admin', 'secretary']):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        from authentication.models import StudentEnrollment
        data = json.loads(request.body)
        enrollment_id = data.get('enrollment_id')
        new_status = data.get('status')
        
        if not enrollment_id or not new_status:
            return JsonResponse({'error': 'ID de matrícula y estado son requeridos'}, status=400)
        
        enrollment = StudentEnrollment.objects.get(id=enrollment_id)
        old_status = enrollment.get_status_display()
        enrollment.status = new_status
        enrollment.save()
        
        # Si se desactiva el estudiante, remover de grupos académicos
        if new_status in ['withdrawn', 'transferred']:
            from authentication.models import GroupMembership
            GroupMembership.objects.filter(
                user=enrollment.student,
                group__group_type='academic',
                is_active=True
            ).update(is_active=False)
        
        return JsonResponse({
            'success': True,
            'message': f'✅ Estado de {enrollment.student.get_full_name()} cambiado de {old_status} a {enrollment.get_status_display()}'
        })
        
    except StudentEnrollment.DoesNotExist:
        return JsonResponse({'error': 'Matrícula no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error al cambiar estado: {str(e)}'}, status=500)


# APIs para la gestión de usuarios desde la interfaz web
@require_http_methods(["POST"])
@login_required
def create_user_api(request):
    """API para crear un nuevo usuario"""
    try:
        if not hasattr(request.user, 'profile') or request.user.profile.role not in ['admin', 'secretary']:
            return JsonResponse({'error': 'No tienes permisos para crear usuarios'}, status=403)
        
        data = json.loads(request.body)
        
        # Validar datos requeridos
        required_fields = ['username', 'first_name', 'last_name', 'role', 'password']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({'error': f'El campo {field} es requerido'}, status=400)
        
        # Verificar que el username no exista
        if User.objects.filter(username=data['username']).exists():
            return JsonResponse({'error': 'El nombre de usuario ya existe'}, status=400)
        
        # Verificar que las contraseñas coincidan
        if data['password'] != data.get('password_confirm'):
            return JsonResponse({'error': 'Las contraseñas no coinciden'}, status=400)
        
        with transaction.atomic():
            # Crear el usuario
            user = User.objects.create_user(
                username=data['username'],
                first_name=data['first_name'],
                last_name=data['last_name'],
                email=data.get('email', ''),
                password=data['password']
            )
            
            # Actualizar el perfil
            user.profile.role = data['role']
            user.profile.phone = data.get('phone', '')
            user.profile.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Usuario {user.get_full_name()} creado exitosamente',
            'user_id': user.id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Error al crear usuario: {str(e)}'}, status=500)


@require_http_methods(["POST"])
@login_required
def toggle_user_status_api(request):
    """API para activar/desactivar usuarios"""
    try:
        if not hasattr(request.user, 'profile') or request.user.profile.role not in ['admin', 'secretary']:
            return JsonResponse({'error': 'No tienes permisos para cambiar estados de usuarios'}, status=403)
        
        data = json.loads(request.body)
        user_id = data.get('user_id')
        activate = data.get('activate', True)
        
        user = User.objects.get(id=user_id)
        
        # No permitir desactivar el propio usuario
        if user == request.user:
            return JsonResponse({'error': 'No puedes desactivar tu propio usuario'}, status=400)
        
        user.is_active = activate
        user.save()
        
        action = 'activado' if activate else 'desactivado'
        return JsonResponse({
            'success': True,
            'message': f'Usuario {user.get_full_name()} {action} exitosamente'
        })
        
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error al cambiar estado: {str(e)}'}, status=500)


@require_http_methods(["GET"])
@login_required
def get_user_details_api(request, user_id):
    """API para obtener detalles de un usuario"""
    try:
        if not hasattr(request.user, 'profile') or request.user.profile.role not in ['admin', 'secretary']:
            return JsonResponse({'error': 'No tienes permisos para ver usuarios'}, status=403)
        
        user = User.objects.select_related('profile').get(id=user_id)
        
        user_data = {
            'id': user.id,
            'username': user.username,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'is_active': user.is_active,
            'date_joined': user.date_joined.strftime('%d/%m/%Y %H:%M'),
            'last_login': user.last_login.strftime('%d/%m/%Y %H:%M') if user.last_login else 'Nunca',
            'profile': {
                'role': user.profile.role,
                'role_display': user.profile.get_role_display(),
                'phone': user.profile.phone,
                'specialty': user.profile.specialty,
                'avatar_url': user.profile.avatar.url if user.profile.avatar else None,
            }
        }
        
        # Si es estudiante, agregar información de matrícula
        if user.profile.role == 'student':
            try:
                enrollment = StudentEnrollment.objects.select_related('homeroom_teacher').get(
                    student=user, academic_year='2025'
                )
                user_data['enrollment'] = {
                    'grade': enrollment.get_grade_display(),
                    'section': enrollment.section,
                    'status': enrollment.get_status_display(),
                    'homeroom_teacher': enrollment.homeroom_teacher.get_full_name() if enrollment.homeroom_teacher else 'Sin asignar',
                    'parent_guardian_name': enrollment.parent_guardian_name,
                    'parent_guardian_phone': enrollment.parent_guardian_phone,
                    'parent_guardian_email': enrollment.parent_guardian_email,
                }
            except StudentEnrollment.DoesNotExist:
                user_data['enrollment'] = None
        
        return JsonResponse(user_data)
        
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error al obtener usuario: {str(e)}'}, status=500)


@require_http_methods(["POST"])
@login_required
def enroll_student_api(request):
    """API para matricular un estudiante"""
    try:
        if not hasattr(request.user, 'profile') or request.user.profile.role not in ['admin', 'secretary']:
            return JsonResponse({'error': 'No tienes permisos para matricular estudiantes'}, status=403)
        
        data = json.loads(request.body)
        
        # Validar datos requeridos
        required_fields = ['student_id', 'grade', 'section', 'academic_year']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({'error': f'El campo {field} es requerido'}, status=400)
        
        student = User.objects.get(id=data['student_id'])
        
        # Verificar que es un estudiante
        if student.profile.role != 'student':
            return JsonResponse({'error': 'Solo se pueden matricular usuarios con rol de estudiante'}, status=400)
        
        # Verificar si ya tiene matrícula para este año
        existing_enrollment = StudentEnrollment.objects.filter(
            student=student,
            academic_year=data['academic_year']
        ).first()
        
        if existing_enrollment:
            return JsonResponse({'error': f'El estudiante ya tiene matrícula para el año {data["academic_year"]}'}, status=400)
        
        with transaction.atomic():
            # Crear la matrícula
            enrollment = StudentEnrollment.objects.create(
                student=student,
                academic_year=data['academic_year'],
                grade=data['grade'],
                section=data['section'],
                status=data.get('status', 'active'),
                homeroom_teacher_id=data.get('homeroom_teacher') if data.get('homeroom_teacher') else None,
                parent_guardian_name=data.get('parent_guardian_name', ''),
                parent_guardian_phone=data.get('parent_guardian_phone', ''),
                parent_guardian_email=data.get('parent_guardian_email', ''),
                created_by=request.user
            )
            
            # Asignar al grupo académico automáticamente
            enrollment.assign_to_academic_group()
        
        return JsonResponse({
            'success': True,
            'message': f'Estudiante {student.get_full_name()} matriculado exitosamente en {enrollment.full_grade}',
            'enrollment_id': enrollment.id
        })
        
    except User.DoesNotExist:
        return JsonResponse({'error': 'Estudiante no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error al matricular estudiante: {str(e)}'}, status=500)


@require_http_methods(["GET"])
@login_required
def export_students_excel_api(request):
    """API para exportar estudiantes a Excel"""
    try:
        if not hasattr(request.user, 'profile') or request.user.profile.role not in ['admin', 'secretary']:
            return JsonResponse({'error': 'No tienes permisos para exportar datos'}, status=403)
        
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        import io
        
        # Crear el libro de trabajo
        wb = Workbook()
        ws = wb.active
        ws.title = "Estudiantes Matriculados"
        
        # Encabezados
        headers = [
            'Nombre Completo', 'Usuario', 'Email', 'Grado', 'Sección', 
            'Estado Matrícula', 'Director de Curso', 'Acudiente', 
            'Teléfono Acudiente', 'Email Acudiente', 'Fecha Matrícula'
        ]
        
        # Estilo para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Obtener datos de estudiantes
        enrollments = StudentEnrollment.objects.select_related(
            'student', 'student__profile', 'homeroom_teacher'
        ).filter(academic_year='2025').order_by('grade', 'section', 'student__first_name')
        
        # Escribir datos
        for row, enrollment in enumerate(enrollments, 2):
            ws.cell(row=row, column=1, value=enrollment.student.get_full_name())
            ws.cell(row=row, column=2, value=enrollment.student.username)
            ws.cell(row=row, column=3, value=enrollment.student.email)
            ws.cell(row=row, column=4, value=enrollment.get_grade_display())
            ws.cell(row=row, column=5, value=enrollment.section)
            ws.cell(row=row, column=6, value=enrollment.get_status_display())
            ws.cell(row=row, column=7, value=enrollment.homeroom_teacher.get_full_name() if enrollment.homeroom_teacher else 'Sin asignar')
            ws.cell(row=row, column=8, value=enrollment.parent_guardian_name)
            ws.cell(row=row, column=9, value=enrollment.parent_guardian_phone)
            ws.cell(row=row, column=10, value=enrollment.parent_guardian_email)
            ws.cell(row=row, column=11, value=enrollment.created_at.strftime('%d/%m/%Y'))
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="estudiantes_matriculados.xlsx"'
        
        # Guardar el archivo en la respuesta
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({'error': f'Error al exportar: {str(e)}'}, status=500)


# ============== APIs para Sistema Académico ==============

@login_required
@require_http_methods(["POST"])
def create_academic_year_api(request):
    """API para crear un nuevo año académico"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'admin':
        return JsonResponse({'error': 'Sin permisos para esta acción'}, status=403)
    
    try:
        from datetime import datetime
        
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        is_current = data.get('is_current', False)
        
        if not all([name, start_date, end_date]):
            return JsonResponse({'error': 'Todos los campos son obligatorios'}, status=400)
        
        # Convertir fechas de string a date objects
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}, status=400)
        
        # Validar que la fecha de inicio sea anterior a la de fin
        if start_date_obj >= end_date_obj:
            return JsonResponse({'error': 'La fecha de inicio debe ser anterior a la fecha de fin'}, status=400)
        
        # Si se marca como actual, desmarcar otros años
        if is_current:
            ExtendedAcademicYear.objects.filter(is_current=True).update(is_current=False)
        
        academic_year = ExtendedAcademicYear.objects.create(
            name=name,
            start_date=start_date_obj,
            end_date=end_date_obj,
            is_current=is_current
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Año académico "{name}" creado exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error al crear año académico: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def update_academic_year_api(request):
    """API para actualizar un año académico"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'admin':
        return JsonResponse({'error': 'Sin permisos para esta acción'}, status=403)
    
    try:
        from datetime import datetime
        
        data = json.loads(request.body)
        year_id = data.get('id')
        name = data.get('name', '').strip()
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        is_current = data.get('is_current', False)
        
        if not year_id:
            return JsonResponse({'error': 'ID del año académico requerido'}, status=400)
        
        if not all([name, start_date, end_date]):
            return JsonResponse({'error': 'Todos los campos son obligatorios'}, status=400)
        
        # Convertir fechas de string a date objects
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}, status=400)
        
        # Validar que la fecha de inicio sea anterior a la de fin
        if start_date_obj >= end_date_obj:
            return JsonResponse({'error': 'La fecha de inicio debe ser anterior a la fecha de fin'}, status=400)
        
        academic_year = get_object_or_404(ExtendedAcademicYear, id=year_id)
        
        # Si se marca como actual, desmarcar otros años
        if is_current and not academic_year.is_current:
            ExtendedAcademicYear.objects.filter(is_current=True).update(is_current=False)
        
        academic_year.name = name
        academic_year.start_date = start_date_obj
        academic_year.end_date = end_date_obj
        academic_year.is_current = is_current
        academic_year.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Año académico "{name}" actualizado exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error al actualizar año académico: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def delete_academic_year_api(request):
    """API para eliminar un año académico"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'admin':
        return JsonResponse({'error': 'Sin permisos para esta acción'}, status=403)
    
    try:
        data = json.loads(request.body)
        year_id = data.get('id')
        
        print(f"DEBUG: Intentando eliminar año académico con ID: {year_id}")  # Debug
        
        if not year_id:
            return JsonResponse({'error': 'ID del año académico requerido'}, status=400)
        
        academic_year = get_object_or_404(ExtendedAcademicYear, id=year_id)
        name = academic_year.name
        
        print(f"DEBUG: Año encontrado: {name}, eliminando...")  # Debug
        
        academic_year.delete()
        
        print(f"DEBUG: Año {name} eliminado exitosamente")  # Debug
        
        return JsonResponse({
            'success': True,
            'message': f'Año académico "{name}" eliminado exitosamente'
        })
        
    except Exception as e:
        print(f"DEBUG: Error al eliminar año académico: {str(e)}")  # Debug
        return JsonResponse({'error': f'Error al eliminar año académico: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def set_current_academic_year_api(request):
    """API para establecer un año académico como actual"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'admin':
        return JsonResponse({'error': 'Sin permisos para esta acción'}, status=403)
    
    try:
        data = json.loads(request.body)
        year_id = data.get('id')
        
        if not year_id:
            return JsonResponse({'error': 'ID del año académico requerido'}, status=400)
        
        # Desmarcar todos los años actuales
        ExtendedAcademicYear.objects.filter(is_current=True).update(is_current=False)
        
        # Marcar el año seleccionado como actual
        academic_year = get_object_or_404(ExtendedAcademicYear, id=year_id)
        academic_year.is_current = True
        academic_year.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Año académico "{academic_year.name}" establecido como actual'
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error al establecer año académico actual: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def create_grade_api(request):
    """API para crear un nuevo grado"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'admin':
        return JsonResponse({'error': 'Sin permisos para esta acción'}, status=403)
    
    try:
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        level = data.get('level', '').strip()
        order = data.get('order', 1)
        
        if not all([name, level]):
            return JsonResponse({'error': 'Nombre y nivel son obligatorios'}, status=400)
        
        grade = Grade.objects.create(
            name=name,
            level=level,
            order=order
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Grado "{name}" creado exitosamente',
            'grade': {
                'id': grade.id,
                'name': grade.name,
                'level': grade.level,
                'order': grade.order
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error al crear grado: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def update_grade_api(request):
    """API para actualizar un grado"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'admin':
        return JsonResponse({'error': 'Sin permisos para esta acción'}, status=403)
    
    try:
        data = json.loads(request.body)
        grade_id = data.get('id')
        name = data.get('name', '').strip()
        level = data.get('level', '').strip()
        order = data.get('order', 1)
        
        if not grade_id:
            return JsonResponse({'error': 'ID del grado requerido'}, status=400)
        
        grade = get_object_or_404(Grade, id=grade_id)
        grade.name = name
        grade.level = level
        grade.order = order
        grade.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Grado "{name}" actualizado exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error al actualizar grado: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def delete_grade_api(request):
    """API para eliminar un grado"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'admin':
        return JsonResponse({'error': 'Sin permisos para esta acción'}, status=403)
    
    try:
        data = json.loads(request.body)
        grade_id = data.get('id')
        
        if not grade_id:
            return JsonResponse({'error': 'ID del grado requerido'}, status=400)
        
        grade = get_object_or_404(Grade, id=grade_id)
        name = grade.name
        grade.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Grado "{name}" eliminado exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error al eliminar grado: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def create_subject_api(request):
    """API para crear una nueva materia"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'admin':
        return JsonResponse({'error': 'Sin permisos para esta acción'}, status=403)
    
    try:
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        area = data.get('area', '').strip()
        code = data.get('code', '').strip()
        description = data.get('description', '').strip()
        
        if not all([name, area]):
            return JsonResponse({'error': 'Nombre y área son obligatorios'}, status=400)
        
        subject = Subject.objects.create(
            name=name,
            area=area,
            code=code,
            description=description
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Materia "{name}" creada exitosamente',
            'subject': {
                'id': subject.id,
                'name': subject.name,
                'area': subject.area,
                'code': subject.code,
                'description': subject.description
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error al crear materia: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def update_subject_api(request):
    """API para actualizar una materia"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'admin':
        return JsonResponse({'error': 'Sin permisos para esta acción'}, status=403)
    
    try:
        data = json.loads(request.body)
        subject_id = data.get('id')
        name = data.get('name', '').strip()
        area = data.get('area', '').strip()
        code = data.get('code', '').strip()
        description = data.get('description', '').strip()
        
        if not subject_id:
            return JsonResponse({'error': 'ID de la materia requerido'}, status=400)
        
        subject = get_object_or_404(Subject, id=subject_id)
        subject.name = name
        subject.area = area
        subject.code = code
        subject.description = description
        subject.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Materia "{name}" actualizada exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error al actualizar materia: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def delete_subject_api(request):
    """API para eliminar una materia"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'admin':
        return JsonResponse({'error': 'Sin permisos para esta acción'}, status=403)
    
    try:
        data = json.loads(request.body)
        subject_id = data.get('id')
        
        if not subject_id:
            return JsonResponse({'error': 'ID de la materia requerido'}, status=400)
        
        subject = get_object_or_404(Subject, id=subject_id)
        name = subject.name
        subject.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Materia "{name}" eliminada exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error al eliminar materia: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def create_course_api(request):
    """API para crear un nuevo curso"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'admin':
        return JsonResponse({'error': 'Sin permisos para esta acción'}, status=403)
    
    try:
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        grade_id = data.get('grade_id')
        academic_year_id = data.get('academic_year_id')
        section = data.get('section', '').strip()
        
        if not all([name, grade_id, academic_year_id]):
            return JsonResponse({'error': 'Nombre, grado y año académico son obligatorios'}, status=400)
        
        grade = get_object_or_404(Grade, id=grade_id)
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        
        course = Course.objects.create(
            name=name,
            grade=grade,
            academic_year=academic_year,
            section=section
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Curso "{name}" creado exitosamente',
            'course': {
                'id': course.id,
                'name': course.name,
                'grade': grade.name,
                'academic_year': academic_year.name,
                'section': course.section
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error al crear curso: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def update_course_api(request):
    """API para actualizar un curso"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'admin':
        return JsonResponse({'error': 'Sin permisos para esta acción'}, status=403)
    
    try:
        data = json.loads(request.body)
        course_id = data.get('id')
        name = data.get('name', '').strip()
        grade_id = data.get('grade_id')
        academic_year_id = data.get('academic_year_id')
        section = data.get('section', '').strip()
        
        if not course_id:
            return JsonResponse({'error': 'ID del curso requerido'}, status=400)
        
        course = get_object_or_404(Course, id=course_id)
        
        if grade_id:
            course.grade = get_object_or_404(Grade, id=grade_id)
        if academic_year_id:
            course.academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        
        course.name = name
        course.section = section
        course.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Curso "{name}" actualizado exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error al actualizar curso: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def delete_course_api(request):
    """API para eliminar un curso"""
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'admin':
        return JsonResponse({'error': 'Sin permisos para esta acción'}, status=403)
    
    try:
        data = json.loads(request.body)
        course_id = data.get('id')
        
        if not course_id:
            return JsonResponse({'error': 'ID del curso requerido'}, status=400)
        
        course = get_object_or_404(Course, id=course_id)
        name = course.name
        course.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Curso "{name}" eliminado exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error al eliminar curso: {str(e)}'}, status=500)